"""
Utilitaires pour l'export de données en Excel et PDF
"""
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from datetime import datetime


class ExcelExporter:
    """Classe pour exporter des données en Excel"""
    
    def __init__(self, filename, sheet_name="Données"):
        self.filename = filename
        self.sheet_name = sheet_name
        self.workbook = Workbook()
        self.worksheet = self.workbook.active
        self.worksheet.title = sheet_name
        self.current_row = 1
        
    def add_title(self, title):
        """Ajouter un titre principal"""
        cell = self.worksheet.cell(row=self.current_row, column=1, value=title)
        cell.font = Font(size=16, bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="1F4788", end_color="1F4788", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        self.current_row += 1
        
    def add_info(self, label, value):
        """Ajouter une ligne d'information"""
        self.worksheet.cell(row=self.current_row, column=1, value=label).font = Font(bold=True)
        self.worksheet.cell(row=self.current_row, column=2, value=value)
        self.current_row += 1
        
    def add_empty_row(self):
        """Ajouter une ligne vide"""
        self.current_row += 1
        
    def add_headers(self, headers):
        """Ajouter les en-têtes de colonnes"""
        for col_num, header in enumerate(headers, 1):
            cell = self.worksheet.cell(row=self.current_row, column=col_num, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        self.current_row += 1
        
    def add_row(self, data, is_total=False):
        """Ajouter une ligne de données"""
        for col_num, value in enumerate(data, 1):
            cell = self.worksheet.cell(row=self.current_row, column=col_num, value=value)
            if is_total:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        self.current_row += 1
        
    def auto_adjust_columns(self):
        """Ajuster automatiquement la largeur des colonnes"""
        for column in self.worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            self.worksheet.column_dimensions[column_letter].width = adjusted_width
            
    def get_response(self):
        """Générer la réponse HTTP"""
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{self.filename}"'
        self.workbook.save(response)
        return response


class PDFExporter:
    """Classe pour exporter des données en PDF"""
    
    def __init__(self, filename, title, orientation='portrait'):
        self.filename = filename
        self.title = title
        self.pagesize = landscape(A4) if orientation == 'landscape' else A4
        self.elements = []
        self.styles = getSampleStyleSheet()
        
        # Style personnalisé pour le titre
        self.title_style = ParagraphStyle(
            'CustomTitle',
            parent=self.styles['Heading1'],
            fontSize=18,
            textColor=colors.HexColor('#1F4788'),
            spaceAfter=30,
            alignment=1  # Center
        )
        
    def add_title(self, title=None):
        """Ajouter un titre"""
        title_text = title or self.title
        self.elements.append(Paragraph(title_text, self.title_style))
        self.elements.append(Spacer(1, 0.5*cm))
        
    def add_info_table(self, data):
        """Ajouter un tableau d'informations"""
        table = Table(data, colWidths=[5*cm, 10*cm])
        table.setStyle(TableStyle([
            ('FONT', (0, 0), (0, -1), 'Helvetica-Bold', 10),
            ('FONT', (1, 0), (1, -1), 'Helvetica', 10),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ]))
        self.elements.append(table)
        self.elements.append(Spacer(1, 0.5*cm))
        
    def add_data_table(self, headers, data, col_widths=None):
        """Ajouter un tableau de données"""
        # Préparer les données avec en-têtes
        table_data = [headers] + data
        
        # Créer le tableau
        if col_widths:
            table = Table(table_data, colWidths=col_widths)
        else:
            table = Table(table_data)
            
        # Style du tableau
        table.setStyle(TableStyle([
            # En-tête
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            
            # Données
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
            ('ALIGN', (0, 1), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            
            # Alternance de couleurs
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
        ]))
        
        self.elements.append(table)
        
    def add_footer_info(self, text):
        """Ajouter une note de bas de page"""
        self.elements.append(Spacer(1, 0.5*cm))
        footer_style = ParagraphStyle(
            'Footer',
            parent=self.styles['Normal'],
            fontSize=8,
            textColor=colors.grey
        )
        self.elements.append(Paragraph(text, footer_style))
        
    def get_response(self):
        """Générer la réponse HTTP"""
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{self.filename}"'
        
        doc = SimpleDocTemplate(
            response,
            pagesize=self.pagesize,
            rightMargin=2*cm,
            leftMargin=2*cm,
            topMargin=2*cm,
            bottomMargin=2*cm
        )
        
        doc.build(self.elements)
        return response


def format_currency(value):
    """Formater une valeur monétaire"""
    try:
        return f"{float(value):,.0f} GNF".replace(',', ' ')
    except:
        return "0 GNF"


def format_date(date_obj):
    """Formater une date"""
    if date_obj:
        return date_obj.strftime('%d/%m/%Y')
    return "N/A"


def format_datetime(datetime_obj):
    """Formater une date et heure"""
    if datetime_obj:
        return datetime_obj.strftime('%d/%m/%Y %H:%M')
    return "N/A"
